import requests
import json
import base64
import openpyxl
import os
from time import sleep

def get_playlist_tracks(playlist_id, file_name):
    
    # Set Filename
    file_name_final = "Tracks_" + file_name + ".xlsx"

    # SECRET INFO FILL IN THESE YOURSELF
    client_id = "1234567890qwertyuiop"
    client_secret = "1234567890qwertyuiop"

    # Obtain an access token
    url = "https://accounts.spotify.com/api/token"
    headers = {
        "Authorization": f"Basic {base64.b64encode(f'{client_id}:{client_secret}'.encode()).decode()}"
    }
    data = {
        "grant_type": "client_credentials"
    }
    response = requests.post(url, headers=headers, data=data)
    access_token = json.loads(response.text)["access_token"]

    # Make API request
    limit = 100
    offset = 0
    url = f"https://api.spotify.com/v1/playlists/{playlist_id}/tracks?limit={limit}&offset={offset}"
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Authorization": f"Bearer {access_token}"
    }

    # Parse Using JSON
    response = requests.get(url, headers=headers)
    data = json.loads(response.text)
    total_tracks = data["total"]

    # Iterate Thru > 100 Tracks
    while offset < total_tracks:
        response = requests.get(url, headers=headers)
        
        if response.status_code != 200:
            #Change Offset, Rerun
            offset += limit
            url = f"https://api.spotify.com/v1/playlists/{playlist_id}/tracks?limit={limit}&offset={offset}"
            continue
        data = json.loads(response.text)
        
        for item in data["items"]:
            try:
                track = item["track"]

                #Write to Excel
                save_track_data(track["name"], track["artists"][0]["name"], track["album"]["name"], file_name_final)

                #print(f"Track: {track_name} | Artist: {artist_name} | Album: {album_name}")

            except (KeyError, TypeError):
                continue

        #Change Offset, Rerun
        offset += limit
        url = f"https://api.spotify.com/v1/playlists/{playlist_id}/tracks?limit={limit}&offset={offset}"
    sleep(10.0)

def save_track_data(track_name, artist_name, album, file_name):

    # Check if the file exists
    if os.path.isfile(file_name):

        # Open the workbook if it exists
        wb = openpyxl.load_workbook(file_name)

        # Select the active worksheet
        ws = wb.active

        # Find the first empty row
        next_row = ws.max_row + 1

        # Write the values to the worksheet
        ws.cell(row=next_row, column=1).value = track_name
        ws.cell(row=next_row, column=2).value = artist_name
        ws.cell(row=next_row, column=3).value = album

        # Save the workbook
        wb.save(file_name)

    else:

        # Create a new workbook if the file doesn't exist
        wb = openpyxl.Workbook()

        # Select the active worksheet
        ws = wb.active

        # Write the headers to the worksheet
        ws.cell(row=1, column=1).value = "Track Name"
        ws.cell(row=1, column=2).value = "Artist Name"
        ws.cell(row=1, column=3).value = "Album"

        # Write the first row of data
        ws.cell(row=2, column=1).value = track_name
        ws.cell(row=2, column=2).value = artist_name
        ws.cell(row=2, column=3).value = album

        # Save the workbook
        wb.save(file_name)



# EXAMPLE USAGE - Parameters are playlist ID and Output Filename
get_playlist_tracks("1234567890qwertyuiop", "FreshTunes")
